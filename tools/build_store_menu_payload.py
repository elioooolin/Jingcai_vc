#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import OrderedDict, defaultdict
from pathlib import Path

from openpyxl import load_workbook


MEAL_TYPE_MAP = {
    "早餐": "breakfast",
    "午餐": "lunch",
    "晚餐": "dinner",
    "高补品": "supplement",
}

MENU_RULES = {
    ("早餐", "菜品"): {"selection_rule": "任选1项", "required_count": 1},
    ("午餐", "菜品"): {"selection_rule": "任选2项", "required_count": 2},
    ("午餐", "汤品"): {"selection_rule": "任选1项", "required_count": 1},
    ("晚餐", "菜品"): {"selection_rule": "任选2项", "required_count": 2},
    ("晚餐", "汤品"): {"selection_rule": "任选1项", "required_count": 1},
    ("高补品", "高补品"): {"selection_rule": "最多可选1项", "required_count": 0},
}

MEAL_ORDER = ["早餐", "午餐", "晚餐", "高补品"]
CATEGORY_ORDER = ["菜品", "汤品", "高补品"]


def normalize_keywords(value):
    if value is None:
        return []
    if isinstance(value, str):
        parts = [part.strip() for part in value.replace("，", ",").replace("、", ",").split(",")]
        return [part for part in parts if part]
    return [str(value).strip()]


def to_nutritional_info(row):
    return {
        "calories": row["热量"],
        "protein": row["蛋白质"],
        "fat": row["脂肪"],
        "carbohydrates": row["碳水化合物"],
    }


def build_payload(xlsx_path: Path, store: str, include_supplement: bool):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(headers, values)) for values in ws.iter_rows(min_row=2, values_only=True)]

    dishes_by_name = OrderedDict()
    menus_by_day = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    supplement_names = []

    for row in rows:
        meal_cn = row.get("餐次")
        category = row.get("类别")
        dish_name = row.get("菜品")
        day = row.get("天数")

        if not meal_cn or not category or not dish_name:
            continue

        if meal_cn == "高补品" and not include_supplement:
            continue

        meal_type = MEAL_TYPE_MAP[meal_cn]
        dish_payload = {
            "name": dish_name,
            "description": row.get("文字介绍") or "",
            "category": category,
            "meal_type": meal_type,
            "ingredients": row.get("食材") or "",
            "keywords": normalize_keywords(row.get("关键词")),
            "chefRecommend": str(row.get("主厨推荐") or "").strip() == "推荐",
            "nutritional_info": to_nutritional_info(row),
            "store": store,
        }

        dishes_by_name.setdefault(dish_name, dish_payload)

        if meal_cn == "高补品":
            supplement_names.append(dish_name)
            continue

        if day is None:
            continue

        menus_by_day[int(day)][meal_cn][category].append(dish_name)

    dishes = list(dishes_by_name.values())
    daily_menus = []

    for day in sorted(menus_by_day):
        meals_payload = OrderedDict()
        for meal_cn in MEAL_ORDER:
            if meal_cn == "高补品":
                continue
            meal_categories = menus_by_day[day].get(meal_cn)
            if not meal_categories:
                continue

            category_payload = OrderedDict()
            for category in CATEGORY_ORDER:
                dish_names = meal_categories.get(category)
                if not dish_names:
                    continue

                rules = MENU_RULES[(meal_cn, category)]
                category_payload[category] = {
                    "selection_rule": rules["selection_rule"],
                    "required_count": rules["required_count"],
                    "dish_names": dish_names,
                }

            if category_payload:
                meals_payload[MEAL_TYPE_MAP[meal_cn]] = category_payload

        daily_menus.append({
            "day": day,
            "store": store,
            "meals": meals_payload,
        })

    return {
        "store": store,
        "source_file": str(xlsx_path),
        "sheet_name": ws.title,
        "include_supplement": include_supplement,
        "dishes": dishes,
        "daily_menus": daily_menus,
        "supplement_names": supplement_names,
    }


def main():
    parser = argparse.ArgumentParser(description="Build CloudBase menu import payloads from store menu xlsx.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--store", required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--include-supplement", action="store_true")
    args = parser.parse_args()

    payload = build_payload(args.xlsx_path, args.store, args.include_supplement)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    dishes_path = args.output_dir / "dishes.json"
    menus_path = args.output_dir / "daily_menus.json"
    meta_path = args.output_dir / "meta.json"

    dishes_path.write_text(json.dumps(payload["dishes"], ensure_ascii=False, indent=2), encoding="utf-8")
    menus_path.write_text(json.dumps(payload["daily_menus"], ensure_ascii=False, indent=2), encoding="utf-8")
    meta_path.write_text(
        json.dumps(
            {
                "store": payload["store"],
                "source_file": payload["source_file"],
                "sheet_name": payload["sheet_name"],
                "include_supplement": payload["include_supplement"],
                "dish_count": len(payload["dishes"]),
                "menu_day_count": len(payload["daily_menus"]),
                "supplement_names": payload["supplement_names"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"wrote {dishes_path}")
    print(f"wrote {menus_path}")
    print(f"wrote {meta_path}")


if __name__ == "__main__":
    main()
